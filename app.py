import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import datetime
from statsmodels.tsa.arima.model import ARIMA
import plotly.express as px
import pymongo
import sys
import openpyxl
import os
import zipfile
from datetime import datetime, timedelta
import io
from CAL import perform
from prophet import Prophet
from prophet.plot import add_changepoints_to_plot
import altair as alt  
from sklearn.metrics import mean_absolute_error, mean_squared_error

UPLOAD_FOLDER = os.path.join(os.getcwd(), "Upload")

df2 = pd.read_csv('weather.csv')
st.set_page_config(page_title="Revenue Forecasting", page_icon=":barchart:", layout="wide")

st.title("Hotel Revenue Forecasting")
# MongoDB connection setup
connection_uri = "mongodb+srv://annu21312:6dPsrXPfhm19YxXl@hello.hes3iy5.mongodb.net/"
client = pymongo.MongoClient(connection_uri, serverSelectionTimeoutMS=30000)
database_name = "Revenue_Forecasting"
db = client[database_name]
collection = db["Forecastin"]

# connection_uri = "mongodb://localhost:27017/"
# client = pymongo.MongoClient(connection_uri)
# database_name = "revenue_database"
# db = client[database_name]
# collection = db["revenue_table1"]

# Retrieve the starting and ending "Business Date" values from the database
pipeline = [
    {"$group": {"_id": None, "minDate": {"$min": "$Business Date"}, "maxDate": {"$max": "$Business Date"}}}
]
result = list(collection.aggregate(pipeline))

if result:
    start_date_str = result[0]["minDate"]
    end_date_str = result[0]["maxDate"]
else:
    # Handle the case where there is no data in the collection
    start_date_str = datetime.now().strftime("%Y-%m-%d")
    end_date_str = datetime.now().strftime("%Y-%m-%d")

# Convert start_date_str and end_date_str to datetime objects
start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
end_date = datetime.strptime(end_date_str, "%Y-%m-%d")


col1, col2=st.columns((2))
with col1:
    date1 = st.date_input("Select Start Date:", start_date)
with col2:
    date2 = st.date_input("Select End Date:", end_date)
# Convert selected dates to ISO format for MongoDB query
date1 = date1.strftime("%Y-%m-%d")
date2 = date2.strftime("%Y-%m-%d")
def home():
    query = {"Business Date": {"$gte": date1, "$lte": date2}}
    # Fetch data from MongoDB
    cursor = collection.find(query)
    # Convert MongoDB cursor to DataFrame
    df = pd.DataFrame(list(cursor))
    if not df.empty:
        # Display 'Room Revenue' for the selected date range
        total_room_revenue = df['Room Revenue'].sum()
        # st.markdown("**Room Revenue Over Time**")
        st.markdown("<div class='section-title'>Room Revenue Over Time</div>", unsafe_allow_html=True)

        # Interactive chart using Plotly with hover info and different colors
        # fig = px.line(df, x='Business Date', y='Room Revenue', title='\nRoom Revenue Over Time\n\n',
        fig = px.line(df, x='Business Date', y='Room Revenue',
                      labels={'Business Date': 'Date', 'Room Revenue': 'Revenue'},
                    #   hover_data={'Room Revenue': ':,.2f'},  # Format hover value with comma separator
                      line_shape='linear',
                      color_discrete_sequence=['#0074D9']
                    #   color_discrete_sequence=px.colors.qualitative.Set1

                      )  

        fig.update_layout(
        xaxis_title_text='Date',
        yaxis_title_text='Revenue',
        xaxis_title_font_size=16,  # Increase the x-axis label font size
        yaxis_title_font_size=16,  # Increase the y-axis label font size
        xaxis_rangeslider_visible=True,
        xaxis_rangeslider_thickness=0.05,  # Adjust the thickness of the range slider
        xaxis_rangeslider_bgcolor='lightgray',  # Set the background color of the range slider
        xaxis_rangeslider_bordercolor='gray',  # Set the border color of the range slider
        xaxis_rangeslider_borderwidth=1,  # Set the border width of the range slider
        xaxis_rangeslider_range=[date1, date2],  # Set the initial date range
        xaxis_rangeselector=dict(
            buttons=list([
                dict(count=7, label="1W", step="day", stepmode="backward"),
                dict(count=1, label="1M", step="month", stepmode="backward"),
                dict(count=3, label="3M", step="month", stepmode="backward"),
                dict(count=6, label="6M", step="month", stepmode="backward"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1Y", step="year", stepmode="backward"),
                dict(step="all")
                ]),
                x=0.01,  # Adjust the horizontal position of the rangeselector
                xanchor='left',  # Anchor the rangeselector to the left
            )
        )
        # st.write(f"Total Room Revenue for the selected date range: ${total_room_revenue:.2f}")

        fig.update_traces(mode="lines+markers", hovertemplate="Date: %{x}<br>Revenue: $%{y:,.2f}<extra></extra>")
        fig.update_xaxes(type='date', showgrid=True, gridwidth=1, gridcolor='lightgray')  # Add gridlines
        fig.update_yaxes(title_text='Revenue', showgrid=True, gridwidth=1, gridcolor='lightgray')  # Add gridlines
       
  
        # Customize the layout
        fig.update_layout(
            margin=dict(l=40, r=20, t=40, b=40),
            xaxis_title='Date',
            yaxis_title='Revenue',
            legend_title='Legend',
            font=dict(family='Arial', size=14),
        )
        st.plotly_chart(fig)
        st.write(f"Total Room Revenue for the selected date range: ${total_room_revenue:.2f}")

        # Button to show/hide the DataFrame
        if 'show_df' not in st.session_state:
            st.session_state.show_df = False

        if st.button("Show Filtered Data"):
            st.session_state.show_df = not st.session_state.show_df

        if st.session_state.show_df:
            # st.write("Filtered Data:")
            st.write(df)

    else:
        st.warning("No data available for the selected date range.")

def display_data():
    # collection = db["Forecasting"]    
    # CSS for styling
    css = """
    body {
        background-color: #ffffff;
        color: #000000;
    }
    .card-container {
        display: flex;
        flex-wrap: wrap;
        justify-content: space-between;
    }

    .card {
        flex-basis: 30%; /* Set card width to 30% for three cards in a row */
        # background-color: #fff;
        padding: 20px;
        margin: 10px 0;
        border-radius: 10px;
        box-shadow: 0px 2px 6px rgba(0, 0, 0, 0.1);
        # transition: transform 0.2s, box-shadow 0.2s;
        transition: transform 0.2s ease-in-out, box-shadow 0.2s ease-in-out;

        color: {card_text_color};
    }
    .card:hover {
        # transform: scale(1.02); /* Add zoom effect on hover */
        transform: scale(1.03);
        box-shadow: 0px 4px 10px rgba(0, 0, 0, 0.2);
    }

    .section-title {
        font-size: 24px;
        font-weight: bold;
        margin-bottom: 20px;
        color: {card_text_color};
    }
    """

    # Apply CSS to the HTML elements
    st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)

    # if section == "Daily Overview":
    st.markdown("<div class='section-title'>Daily Overview</div>", unsafe_allow_html=True)
    # MongoDB query to find data for the two selected dates
    query = {"Business Date": {"$in": [date1, date2]}}
    cursor = collection.find(query)
    
    # Convert MongoDB cursor to DataFrame
    df = pd.DataFrame(list(cursor))
    
    if not df.empty:
        # Extract data for the two selected dates
        date1_data = df[df["Business Date"] == date1]
        date2_data = df[df["Business Date"] == date2]
        # Check if data exists for both selected dates
        if not date1_data.empty and not date2_data.empty:

            # Function to display values for a given field
            def display_field(field_name, date1_data, date2_data):
                # Calculate the difference between the two dates
                field_diff = date2_data[field_name].iloc[0] - date1_data[field_name].iloc[0]
                
                # Choose a color based on the field_diff (you can customize this logic)
                color = "red" if isinstance(field_diff, (int, float)) and field_diff < 0 else "green"
                
                # Display the values in a modern design button
                st.markdown(
                    f"<div class='card'>"
                    # f"<h5>{field_name}</h5>"
                    f"<p>{field_name}</p>"

                    f"<div class='card-values'>"
                    # f"<span style='font-size: 28px;text-align: left'>{date1_data[field_name].iloc[0]}&emsp;</span> <span class='red-text' style='color: {color};text-align: center'>{field_diff:.2f}</span> &emsp;<span style='text-align: right;'>{date2_data[field_name].iloc[0]}</span>"
                    f"<div class='card-values' style='display: flex; justify-content: space-between; align-items: center;'>"
                    # f"<div style='text-align: left; flex-basis: 33%;font-size: 28px;color: {color};'>{date1_data[field_name].iloc[0]}</div>"
                    f"<div style='text-align: left; flex-basis: 33%; color: {color};'>{date1_data[field_name].iloc[0]}</div>"

                    f"<div style='text-align: center; flex-basis: 33%; font-weight: bold;'>{field_diff:.2f}</div>"
                    f"<div style='text-align: right; flex-basis: 33%;'>{date2_data[field_name].iloc[0]}</div>"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
            # Display values for each field
            fields = ["Occupancy %", "ARR", "Arrival Rooms", "OOO Rooms", "Pax", "Room Revenue", "Rooms for Sale", "Departure Rooms", "House Use", "Total Room Inventory"]
            col1, col2, col3 = st.columns(3)

            # # Display values for each field
            for i, field in enumerate(fields):
                if i % 3 == 0:
                    card_column = col1
                elif i % 3 == 1:
                    card_column = col2
                else:
                    card_column = col3
                with card_column:
                    display_field(field, date1_data, date2_data)

        else:
            st.warning("Data not available for one or both of the selected dates.")
            
def  Revenue_analysis():
    query = {"Business Date": {"$gte": date1, "$lte": date2}}
    # Fetch data from MongoDB
    cursor = collection.find(query)
    # Convert MongoDB cursor to DataFrame
    df = pd.DataFrame(list(cursor))
    st.write(df.columns)

    # data = list(collection.find())  # Fetch data as a list of documents
    # df = pd.DataFrame(data)
    # Define available chart options
    chart_options = ["Revenue Breakdown", "Group Bookings", "Demand and Supply", "Customer Segmentation", "Ooo Rooms", "Inclusion Revenue", "Total Room Inventory"]
    selected_option = chart_options[0]
    selected_option = st.selectbox("Select Chart Option", chart_options, format_func=lambda option: option)
    df['Business Date'] = pd.to_datetime(df['Business Date'])
    fig = None

    if selected_option == "Revenue Breakdown":
        selected_columns = ["Room Revenue", "Individual Revenue", "Confirmed Group Revenue", "Tentative Group Revenue"]
        fig = px.bar(df, x="Business Date", y=selected_columns, barmode="stack", title="Revenue Breakdown")
    elif selected_option == "Group Bookings":
        selected_columns = ["Confirmed Group Revenue", "Tentative Group Revenue"]
        fig = px.line(df, x="Business Date", y=selected_columns, title="Group Booking Trends")
    elif selected_option == "Demand and Supply":
        selected_columns = ["Arrival Rooms", "Compliment Rooms"]
        fig = px.bar(df, x="Business Date", y=selected_columns, title="Supply and Demand")
    elif selected_option == "Customer Segmentation":
        selected_columns = ["Individual Revenue", "Confirmed Group Revenue"]
        fig = px.line(df, x="Business Date", y=selected_columns, title="Customer Segmentation")
    elif selected_option == "Ooo Rooms": #Out of Order (OOO) Rooms
        selected_columns = ["OOO Rooms"]
        fig = px.bar(df, x="Business Date", y=selected_columns, title="Out of Order (OOO) Rooms")
    elif selected_option == "Inclusion Revenue":
        selected_columns = ["Inclusion Revenue"]
        fig = px.line(df, x="Business Date", y=selected_columns, title="Inclusion Revenue")
    else:  # "Total Room Inventory"
        selected_columns = ["Total Room Inventory"]
        fig = px.bar(df, x="Business Date", y=selected_columns, title="Total Room Inventory")

    # Customize the chart's appearance
    fig.update_layout(
        xaxis_title="Business Date",
        yaxis_title="Value",
        template="plotly_dark",  # Apply a dark theme for aesthetics
        # xaxis_rangeselector=dict(
        #     buttons=list([
        #         dict(count=7, label="Revenue Breakdown"),
        #         dict(count=1, label="Group Bookings"),
        #         dict(count=3, label="Demand and Supply"),
        #         dict(count=6, label="Customer Segmentation"),
        #         dict(count=1, label="Ooo Rooms"),
        #         dict(count=1, label="Inclusion Revenue"),
        #         dict(count=1, label="Total Room Inventory"),
        #         dict(step="all")
        #         ]),
        #         # x=0.01,  # Adjust the horizontal position of the rangeselector
        #         # xanchor='left',  # Anchor the rangeselector to the left
        #     )
        xaxis_rangeselector=dict(
            buttons=list([
                dict(count=7, label="1W", step="day"),
                dict(count=1, label="1M", step="month"),
                dict(count=3, label="3M", step="month"),
                dict(count=6, label="6M", step="month"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1Y", step="year"),
                dict(step="all")
                ]),
                x=0.01,  # Adjust the horizontal position of the rangeselector
                xanchor='left',  # Anchor the rangeselector to the left
            )
    )

    # Add interactivity with hover information
    fig.update_traces(hoverinfo='y+name')

    # Display the interactive chart
    st.plotly_chart(fig)

    # Provide an option to download the chart as an image
    if st.button("Download Chart as Image"):
        st.image(fig.to_image(format="png"), use_container_width=True)

def report():
    # Create the upload folder if it doesn't exist
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    # Add custom CSS styles
    st.markdown(
        """
        <style>
        body {
            font-family: Arial, sans-serif;
            background-color: #f4f4f4;
            margin: 0;
            padding: 0;
        }

        .container {
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            background-color: #ffffff;
            border-radius: 5px;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);
        }

        .header {
            text-align: center;
            font-size: 24px;
            margin-bottom: 20px;
        }

        .input-section {
            margin-bottom: 20px;
        }

        .button-section {
            text-align: center;
        }

        .download-section {
            margin-top: 20px;
            text-align: center;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    # Main content area
    st.markdown("<div class='section-title'>Forecast Report</div>", unsafe_allow_html=True)

    # User input section
    st.markdown('<div class="input-section">', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        f1 = st.file_uploader("Upload File 1 (f1.xlsx)", type=["xlsx"])
    with col2:
        f2 = st.file_uploader("Upload File 2 (f2.xlsx)", type=["xlsx"])
    group_confirm = st.text_input("Group Confirm Number")
    st.markdown('</div>', unsafe_allow_html=True)

    # Process Data button
    st.markdown('<div class="button-section">', unsafe_allow_html=True)
    if st.button("Process Data"):
        if f1 and f2 and group_confirm:
            # Save uploaded files to the upload folder
            with open(os.path.join(UPLOAD_FOLDER, "f1.xlsx"), "wb") as f1_file:
                f1_file.write(f1.read())
            with open(os.path.join(UPLOAD_FOLDER, "f2.xlsx"), "wb") as f2_file:
                f2_file.write(f2.read())

            # # Load Excel files and sheets
            excel_file1 = openpyxl.load_workbook(os.path.join(UPLOAD_FOLDER, "f1.xlsx"))
            excel_sheet1 = excel_file1["Day on Day FC"]
            excel_sheet3 = excel_file1["Revenue Summary"]
            excel_sheet4 = excel_file1["Segment_Summary"]

            excel_file2 = openpyxl.load_workbook(os.path.join(UPLOAD_FOLDER, "f2.xlsx"))
            excel_sheet2 = excel_file2['History and Forecast Report']

            # Code for manually adding group confirm number
            excel_sheet2.cell(row=5, column=11).value = group_confirm
            perform(excel_file1,excel_sheet1, excel_sheet2,excel_sheet3,excel_sheet4)
            excel_file1.save(os.path.join(UPLOAD_FOLDER, "Final_Report.xlsx"))
            # Create a ZIP archive
            with zipfile.ZipFile(os.path.join(UPLOAD_FOLDER, "output.zip"), 'w') as zip_file:
                # Add each file to the archive
                for file_path in [os.path.join(UPLOAD_FOLDER, "f1.xlsx"), os.path.join(UPLOAD_FOLDER, "Final_Report.xlsx")]:
                    zip_file.write(file_path)

            # Download the ZIP archive
            with open(os.path.join(UPLOAD_FOLDER, "output.zip"), "rb") as zip_file:
                zip_contents = zip_file.read()
            st.markdown("### Download Results")
            st.download_button("Download Output ZIP", data=zip_contents, file_name="output.zip")
    st.markdown('</div>', unsafe_allow_html=True)

    # Close the container div
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
def prophet():
    database_name = "Revenue_Forecasting"
    db = client[database_name]

    # Fetch data from MongoDB
    collection4 = db["Prophet"]
    cursor2 = collection4.find({})
    data = pd.DataFrame(list(cursor2))

    data.rename(columns={'Date': 'ds', 'Revenue': 'y'}, inplace=True)

    # Initialize and fit the Prophet model
    model = Prophet()
    model.fit(data)

    # Create a future DataFrame for forecasting (forecast for the next 7 days)
    future = model.make_future_dataframe(periods=7)

    # Generate forecasts
    forecast = model.predict(future)

    # ACTUAL REVENUE FOR 7 DAYS
    actual_data = data.tail(7)

    # PREDICTIONS FOR 7 DAYS
    next_7_days_forecast = forecast.tail(7)

    predictions_for_7_days = [int(x) for x in list(next_7_days_forecast['yhat'])]
    actual_data_for_7_days = [int(x) for x in list(actual_data['y'])]

    # Calculate accuracy as a percentage
    accuracy = np.mean([pred / actual * 100 for pred, actual in zip(next_7_days_forecast['yhat'], actual_data['y'])])
    # Calculate accuracy metrics
    mae = mean_absolute_error(actual_data['y'], next_7_days_forecast['yhat'])
    mse = mean_squared_error(actual_data['y'], next_7_days_forecast['yhat'])
    rmse = np.sqrt(mse)

    # trend = forecast[['ds', 'trend']]
    # seasonality = forecast[['ds', 'seasonal', 'seasonal_lower', 'seasonal_upper']]
    # holidays = forecast[['ds', 'holidays']]

    # Page Title and Header
    # st.markdown('Forecasting for next week')
    st.markdown("<div class='section-title'>Forecasting for next week</div>", unsafe_allow_html=True)

    # Data Presentation
    col1, col2=st.columns((2))
    with col1:
        st.markdown('Prediction')
        st.dataframe(pd.DataFrame({'Date': next_7_days_forecast['ds'], 'Predicted Revenue': predictions_for_7_days}))
    with col2:
        st.markdown('Actual Revenue')
        st.dataframe(pd.DataFrame({'Date': actual_data['ds'], 'Actual Revenue': actual_data_for_7_days}))

    # CSS for styling
    st.markdown(
        """
        <style>
        /* Adjust text size */
        .big-text {
            font-size: 20px;
        }
        
        .small-text {
            font-size: 14px; /* Adjust the font size as needed */
        }

        /* Style the table */
        table.dataframe {
            font-size: 16px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # st.subheader(f'Accuracy: {accuracy:.2f}%')
    st.markdown(f'Accuracy: {accuracy:.2f}%')
    # st.markdown("<style>.small-text { font-size: 14px; }</style>", unsafe_allow_html=True)

    # # Visualize the forecast using Altair
    # st.header('Revenue Forecast Plot (Altair Chart):')
    # # Create an Altair Chart from the Matplotlib plot
    # forecast_plot = alt.Chart(forecast).mark_line(point=True).encode(
    #     x='ds:T',
    #     y='Revenue Prediction:Q',
    # ).interactive()  # Make the chart interactive
    # st.altair_chart(forecast_plot, use_container_width=True)

    # Create a styled Altair chart
    st.markdown("<div class='section-title'>Revenue Forecast Plot</div>", unsafe_allow_html=True)
    # st.header('Revenue Forecast Plot (Altair Chart):')
    forecast_chart = alt.Chart(forecast).mark_line().encode(
        x=alt.X('ds:T', title='Date', axis=alt.Axis(labelFontSize=12, titleFontSize=14)),
        y=alt.Y('yhat:Q', title='Revenue Prediction', axis=alt.Axis(labelFontSize=12, titleFontSize=14)),
        color=alt.value('blue'),  # Line color
        strokeWidth=alt.value(2),  # Line width
    ).properties(
        width=700,  # Chart width
        height=400,  # Chart height
        title='Revenue Forecast',  # Chart title
    ).configure_title(
        fontSize=16,  # Title font size
        anchor='start'  # Title alignment
    ).configure_axis(
        labelFontSize=12,  # Axis label font size
        titleFontSize=14  # Axis title font size
    )
    # Display the Altair chart in Streamlit
    st.altair_chart(forecast_chart, use_container_width=True)
    # -------------------------------
    st.markdown("<div class='section-title'>Revenue Components Plot</div>", unsafe_allow_html=True)
    # st.header('Revenue Components Plot (Altair Chart):')
    # Create an Altair Chart for components that closely resembles the Matplotlib plot
    components_chart = alt.Chart(forecast).mark_area().encode(
        x='ds:T',
        y='trend:Q',
        y2='yhat_upper:Q',
        opacity=alt.value(0.5)
    ).interactive()  # Make the chart interactive

    # Display the Altair chart in Streamlit
    st.altair_chart(components_chart, use_container_width=True)

    # ------------------------------
    # # Create a styled Altair chart for trend component
    # st.header('Trend Component (Altair Chart):')
    # st.markdown("<div class='section-title'>Revenue Forecast Plot</div>", unsafe_allow_html=True)
    trend_chart = alt.Chart(forecast).mark_line().encode(
        x=alt.X('ds:T', title='Date', axis=alt.Axis(labelFontSize=12, titleFontSize=14)),
        y=alt.Y('trend:Q', title='Trend', axis=alt.Axis(labelFontSize=12, titleFontSize=14)),
        color=alt.value('blue'),  # Line color
        strokeWidth=alt.value(2),  # Line width
    ).properties(
        width=700,  # Chart width
        height=400,  # Chart height
        title='Trend Component',  # Chart title
    ).configure_title(
        fontSize=16,  # Title font size
        anchor='start'  # Title alignment
    ).configure_axis(
        labelFontSize=12,  # Axis label font size
        titleFontSize=14  # Axis title font size
    )

    # Display the Trend Component chart in Streamlit
    st.altair_chart(trend_chart, use_container_width=True)

# Home, Daily_Overview, Report, Performance, Future_Months_and_Pickup=st.tabs(["Home", "Daily Overview","Report", "Performance", "Future Months and Pickup"])
Home, Daily_Overview,Revenue_Analysis, Report, Pridiction=st.tabs(["Home", "Daily Overview", "Revenue Analysis","Report", "Pridiction"])

with Home:
    home()
with Daily_Overview:
    display_data()
with  Revenue_Analysis:
    Revenue_analysis()
with Report:
    report()
with Pridiction:
    prophet()
client.close()


