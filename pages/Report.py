import streamlit as st
import pandas as pd
import openpyxl
import os
import zipfile
import io
from CAL import perform
# import Convert_toSQL as cts
# Set the upload folder
UPLOAD_FOLDER = os.path.join(os.getcwd(), "Upload")
st.set_page_config(page_title="Revenue Forecasting", page_icon=":barchart:", layout="wide")

# Create the upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# Add custom CSS styles
st.markdown(
    """
    <style>
    body {
        font-family: Arial, sans-serif;
        background-color: #f4f4f4;
        margin: 0;
        padding: 0;
    }

    .container {
        max-width: 800px;
        margin: 0 auto;
        padding: 20px;
        background-color: #ffffff;
        border-radius: 5px;
        box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);
    }

    .header {
        text-align: center;
        font-size: 24px;
        margin-bottom: 20px;
    }

    .input-section {
        margin-bottom: 20px;
    }

    .button-section {
        text-align: center;
    }

    .download-section {
        margin-top: 20px;
        text-align: center;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
# Main content area
# st.markdown('<div class="container">', unsafe_allow_html=True)
# st.markdown('<div class="header">Revenue Forecasting App</div>', unsafe_allow_html=True)
st.title('Forecast Report')

# User input section
st.markdown('<div class="input-section">', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    f1 = st.file_uploader("Upload File 1 (f1.xlsx)", type=["xlsx"])
with col2:
    f2 = st.file_uploader("Upload File 2 (f2.xlsx)", type=["xlsx"])
group_confirm = st.text_input("Group Confirm Number")
st.markdown('</div>', unsafe_allow_html=True)

# Process Data button
st.markdown('<div class="button-section">', unsafe_allow_html=True)
if st.button("Process Data"):
    if f1 and f2 and group_confirm:
        # Save uploaded files to the upload folder
        with open(os.path.join(UPLOAD_FOLDER, "f1.xlsx"), "wb") as f1_file:
            f1_file.write(f1.read())
        with open(os.path.join(UPLOAD_FOLDER, "f2.xlsx"), "wb") as f2_file:
            f2_file.write(f2.read())

        # # Load Excel files and sheets
        excel_file1 = openpyxl.load_workbook(os.path.join(UPLOAD_FOLDER, "f1.xlsx"))
        excel_sheet1 = excel_file1["Day on Day FC"]
        excel_sheet3 = excel_file1["Revenue Summary"]
        excel_sheet4 = excel_file1["Segment_Summary"]

        excel_file2 = openpyxl.load_workbook(os.path.join(UPLOAD_FOLDER, "f2.xlsx"))
        excel_sheet2 = excel_file2['History and Forecast Report']

        # Code for manually adding group confirm number
        excel_sheet2.cell(row=5, column=11).value = group_confirm
        perform(excel_file1,excel_sheet1, excel_sheet2,excel_sheet3,excel_sheet4)
        excel_file1.save(os.path.join(UPLOAD_FOLDER, "Final_Report.xlsx"))
        # Create a ZIP archive
        with zipfile.ZipFile(os.path.join(UPLOAD_FOLDER, "output.zip"), 'w') as zip_file:
            # Add each file to the archive
            for file_path in [os.path.join(UPLOAD_FOLDER, "f1.xlsx"), os.path.join(UPLOAD_FOLDER, "Final_Report.xlsx")]:
                zip_file.write(file_path)

        # Download the ZIP archive
        with open(os.path.join(UPLOAD_FOLDER, "output.zip"), "rb") as zip_file:
            zip_contents = zip_file.read()
        st.markdown("### Download Results")
        st.download_button("Download Output ZIP", data=zip_contents, file_name="output.zip")
st.markdown('</div>', unsafe_allow_html=True)

# Close the container div
st.markdown('</div>', unsafe_allow_html=True)

# Display instructions
# st.markdown('<div class="container">', unsafe_allow_html=True)
# st.write("Upload your files and enter the Group Confirm Number in the sidebar.")
st.markdown('</div>', unsafe_allow_html=True)
