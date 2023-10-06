import pandas as pd
from sqlalchemy import create_engine
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
import datetime as dt
import os

#4files


# create an SQLAlchemy engine
engine = create_engine('sqlite:///mydatabase.db', echo=True)

# read the Excel file into a pandas dataframe
path = r"D:\Arjun_Sir-IP-master\History and Forecast/"  
# path = "/content/drive/My Drive/History and Forecast"

columns = ["Actual Date", "Business Date", "Rooms Sold", "Rooms for Sale", "Arrival Rooms", "Compliment Rooms", "House Use", "Hold", "Individual Confirm", "Individual Tentative", "Group Confirm", "Group Tentative", "Occupancy %", "Room Revenue", "ARR", "Inclusion Revenue", "Departure Rooms", "OOO Rooms", "Pax", "Individual Revenue", "Individual ARR", "Confirmed Group Revenue", "Confirmed Group ARR", "Tentative Group Revenue", "Tentative Group ARR", "Total Room Inventory"]
df = pd.DataFrame(columns=columns)

for filename in os.listdir(path):
  if filename.endswith(".xlsx"):
    filepath = os.path.join(path, filename)
    wb = load_workbook(filename=filepath, read_only=True)
    ws = wb.active
    
    actual_date_str = ws.cell(row=8, column=1).value
    actual_date_str = datetime.now()
    actual_date = actual_date_str.strftime("%Y-%m-%d")

    for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=26, values_only=True):
      if isinstance(row[0], (datetime, pd.Timestamp)):
        business_date = row[0].strftime("%Y-%m-%d")
        row_dict = {"Actual Date": actual_date, "Business Date": business_date}

        # Add the data from columns C to Z to the dictionary
        for i in range(2, 26):
          row_dict[columns[i]] = row[i]

        # Append the row data to the dataframe
        df = pd.concat([df, pd.DataFrame([row_dict])], ignore_index=True)

# Print the resulting dataframe
#print(df)
# write the dataframe to the SQLAlchemy database
#df.to_sql('mytable', con=engine, if_exists='replace', index=False)

import pymongo
host = "localhost"
port = 27017
username = "Annu"
database_name = "Revenue_Forecasting"
# database_name = "Forecasting"

connection_uri = f"mongodb+srv://annu21312:6dPsrXPfhm19YxXl@hello.hes3iy5.mongodb.net/"
client = pymongo.MongoClient(connection_uri)
db = client[database_name]
# collection = db["revenue_table1"]

collection = db["Forecasting"]
collection.delete_many({})
data_to_insert = df.to_dict(orient='records')
result = collection.insert_many(data_to_insert)

collection1 = db["History"]
collection1.delete_many({})
df1 = pd.read_excel("History and Forecast Report-20230208.xlsx")
data1 = df1.to_dict(orient="records")
result1 = collection1.insert_many(data1)

collection2 = db["History_Fore"]
collection2.delete_many({})
df2 = pd.read_excel("History and Forecast Report-20220401-20230205.xls")
data2 = df2.to_dict(orient="records")
result2 = collection2.insert_many(data2)

collection3 = db["Covid"]
collection3.delete_many({})
df3 = pd.read_excel("covid_room_revenue.xlsx")
value = "1"
df3.fillna(value, inplace=True)
data3 = df3.to_dict(orient="records")
result2 = collection3.insert_many(data3)

client.close()
